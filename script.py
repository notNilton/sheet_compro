import csv
import re
import pandas as pd
import logging
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter

# Configurando logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')

def ler_csv(file_path, delimiter=';', encoding='utf-8-sig'):
    try:
        logging.info(f"Lendo o arquivo CSV: {file_path}")
        return pd.read_csv(file_path, delimiter=delimiter, encoding=encoding)
    except FileNotFoundError:
        logging.error(f"Erro: Arquivo '{file_path}' não encontrado.")
        return None

def escrever_csv(df, file_path, delimiter=';', encoding='utf-8-sig'):
    logging.info(f"Escrevendo o arquivo CSV: {file_path}")
    df.to_csv(file_path, sep=delimiter, encoding=encoding, index=False)

def comparar_arquivos_csv(arquivo1, coluna1, arquivo2, coluna2, arquivo_saida):
    df1 = ler_csv(arquivo1)
    df2 = ler_csv(arquivo2)
    
    if df1 is None or df2 is None:
        return

    # Verificação de colunas
    if coluna1 not in df1.columns or coluna2 not in df2.columns:
        logging.error(f"Erro: Coluna '{coluna1}' ou '{coluna2}' não encontrada.")
        return

    # Fazer a junção (merge) dos dois DataFrames
    resultado = pd.merge(df2, df1, how='inner', left_on=coluna2, right_on=coluna1)
    escrever_csv(resultado, arquivo_saida)
    logging.info(f"Linhas correspondentes escritas em '{arquivo_saida}'")
    
    csv_para_xlsx(arquivo_saida)

def remover_valores(file_path, keywords, column_name):
    df = ler_csv(file_path)
    
    if df is None:
        return

    if column_name not in df.columns:
        logging.error(f"Erro: Coluna '{column_name}' não encontrada em '{file_path}'")
        return

    pattern = '|'.join([rf'\b{keyword}\b' for keyword in keywords])
    df = df[~df[column_name].str.contains(pattern, case=False, na=False)]
    escrever_csv(df, file_path)
    logging.info(f"Linhas contendo qualquer uma das palavras {keywords} na coluna '{column_name}' foram removidas de '{file_path}'")

def csv_para_xlsx(arquivo_csv):
    try:
        df = pd.read_csv(arquivo_csv, delimiter=';')
        arquivo_xlsx = arquivo_csv.replace('.csv', '.xlsx')
        df.to_excel(arquivo_xlsx, index=False)
        logging.info(f"Arquivo Excel salvo como '{arquivo_xlsx}'")
    except FileNotFoundError:
        logging.error(f"Erro: Arquivo '{arquivo_csv}' não encontrado.")

def ordenar_xlsx_por_coluna(arquivo_xlsx, nome_coluna):
    try:
        wb = load_workbook(filename=arquivo_xlsx)
        ws = wb.active

        col_idx = None
        for cell in ws[1]:
            if cell.value == nome_coluna:
                col_idx = cell.column

        if col_idx is None:
            logging.error(f"Erro: Coluna '{nome_coluna}' não encontrada em '{arquivo_xlsx}'")
            return

        data = list(ws.iter_rows(min_row=2, values_only=True))
        data.sort(key=lambda row: row[col_idx-1])

        for row_idx, row_data in enumerate(data, start=2):
            for col_idx, value in enumerate(row_data, start=1):
                ws.cell(row=row_idx, column=col_idx, value=value)

        wb.save(arquivo_xlsx)
        logging.info(f"Linhas ordenadas por '{nome_coluna}' em '{arquivo_xlsx}'")
    
    except FileNotFoundError:
        logging.error(f"Erro: Arquivo '{arquivo_xlsx}' não encontrado.")

# Exemplo de uso:
arquivo1 = './planilhas_tratadas/processos.csv'
arquivo2 = './planilhas_tratadas/relacao_servidores.csv'

arquivo_saida = './planilhas_resultados/processos_servidores.csv'
arquivo_xlsx_saida = './planilhas_resultados/processos_servidores.xlsx'

comparar_arquivos_csv(arquivo1, 'NOME_PARTE_ATIVA', arquivo2, 'NOME', arquivo_saida)
ordenar_xlsx_por_coluna(arquivo_xlsx_saida, 'NOME')